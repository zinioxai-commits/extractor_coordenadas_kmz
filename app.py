import streamlit as st
import zipfile
import re
import io
import json
import time
import pandas as pd
import urllib.request
from pyproj import Transformer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Redsetel · Extractor KMZ",
    page_icon="📡",
    layout="centered"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Barlow:wght@300;400;600;700&family=Barlow+Condensed:wght@600;700&display=swap');
html, body, [class*="css"] { font-family: 'Barlow', sans-serif; }
.stApp { background-color: #f4f6f9; color: #1a1a2e; }
.topbar {
    background: linear-gradient(135deg, #0D2650 0%, #1a3a6e 100%);
    padding: 1.2rem 2rem; border-radius: 12px;
    display: flex; align-items: center; gap: 1.5rem;
    margin-bottom: 1.8rem;
    box-shadow: 0 4px 20px rgba(13,38,80,0.25);
    border-bottom: 3px solid #CC1E27;
}
.topbar .brand {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 2rem; font-weight: 700; letter-spacing: 1px; line-height: 1;
}
.topbar .brand span.red   { color: #CC1E27; }
.topbar .brand span.white { color: #FFFFFF; }
.topbar .subtitle { color: rgba(255,255,255,0.55); font-size: 0.82rem; font-weight: 300; margin-top: 0.15rem; }
.topbar .divider  { width: 2px; height: 42px; background: rgba(255,255,255,0.15); border-radius: 2px; }
.topbar .tool-title { color: #FFFFFF; font-size: 1rem; font-weight: 600; line-height: 1.2; }
.topbar .tool-desc  { color: rgba(255,255,255,0.45); font-size: 0.78rem; font-weight: 300; margin-top: 0.1rem; }
.card {
    background: #FFFFFF; border-radius: 10px; padding: 1.5rem;
    margin-bottom: 1.2rem; box-shadow: 0 2px 10px rgba(0,0,0,0.06);
    border-top: 3px solid #CC1E27;
}
.card-blue { border-top-color: #0D2650; }
.section-label {
    font-family: 'Barlow Condensed', sans-serif; font-size: 0.72rem; font-weight: 700;
    color: #0D2650; letter-spacing: 1.5px; text-transform: uppercase; margin-bottom: 0.6rem;
}
.metric-row { display: flex; gap: 0.8rem; margin: 0.5rem 0; flex-wrap: wrap; }
.metric-card {
    flex: 1; min-width: 80px;
    background: #f8f9fc; border: 1px solid #e4e8f0;
    border-radius: 8px; padding: 0.9rem 1rem;
    border-left: 3px solid #CC1E27; text-align: center;
}
.metric-card.blue { border-left-color: #0D2650; }
.metric-card .val {
    font-family: 'Barlow Condensed', sans-serif; font-size: 2rem;
    font-weight: 700; color: #CC1E27; line-height: 1;
}
.metric-card.blue .val { color: #0D2650; }
.metric-card .lbl { font-size: 0.7rem; color: #888; margin-top: 0.25rem; text-transform: uppercase; letter-spacing: 0.5px; font-weight: 600; }
[data-testid="stFileUploader"] {
    border: 2px dashed #c8d0e0 !important; border-radius: 10px !important; background: #f8f9fc !important;
}
[data-testid="stFileUploader"]:hover { border-color: #CC1E27 !important; }
[data-testid="stSelectbox"] > div > div {
    background: #f8f9fc !important; border-color: #c8d0e0 !important;
    color: #0D2650 !important; font-weight: 600 !important; border-radius: 8px !important;
}
[data-testid="stDownloadButton"] button {
    background: linear-gradient(135deg, #CC1E27 0%, #a8151d 100%) !important;
    color: #FFFFFF !important; font-family: 'Barlow Condensed', sans-serif !important;
    font-weight: 700 !important; font-size: 1rem !important; letter-spacing: 1px !important;
    border: none !important; border-radius: 8px !important; width: 100% !important;
    padding: 0.7rem 2rem !important; box-shadow: 0 4px 14px rgba(204,30,39,0.3) !important;
}
.info-box {
    background: #eef4ff; border-left: 3px solid #0D2650;
    border-radius: 6px; padding: 0.7rem 1rem;
    font-size: 0.82rem; color: #0D2650; margin-bottom: 0.8rem;
}
.badge {
    display: inline-block; background: #0D2650; color: white;
    font-size: 0.7rem; font-weight: 700; padding: 0.2rem 0.6rem;
    border-radius: 4px; letter-spacing: 0.5px; margin-left: 0.5rem;
}
.empty-state { text-align: center; padding: 3rem 1rem; color: #bbc4d4; font-size: 0.9rem; }
.empty-state .icon { font-size: 2.5rem; margin-bottom: 0.5rem; }
.footer {
    text-align: center; padding: 1.2rem; margin-top: 2rem;
    background: #0D2650; border-radius: 8px;
    color: rgba(255,255,255,0.4); font-size: 0.75rem; letter-spacing: 0.3px;
}
.footer strong { color: rgba(255,255,255,0.7); }
</style>
""", unsafe_allow_html=True)


# ── Funciones KMZ ─────────────────────────────────────────────────
def leer_kml(archivo_bytes: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(archivo_bytes)) as kmz:
        kml_file = next(f for f in kmz.namelist() if f.endswith(".kml"))
        return kmz.read(kml_file).decode("utf-8")


def detectar_formato(kml_texto: str) -> str:
    """
    Detecta el formato del KMZ:
    - 'mwm'    : usa mwm:customName (Maps.me / OruxMaps)
    - 'simple' : nombre completo en <name>, altura en <Camera>
    """
    if "mwm:customName" in kml_texto:
        return "mwm"
    return "simple"


def extraer_puntos(kml_texto: str) -> tuple[list[dict], str]:
    fmt        = detectar_formato(kml_texto)
    placemarks = re.findall(r"<Placemark>(.*?)</Placemark>", kml_texto, re.DOTALL)
    puntos     = []

    for i, pm in enumerate(placemarks):
        # Solo procesar puntos (ignorar LineString, Polygon, etc.)
        if "<Point>" not in pm:
            continue

        coords = re.search(r"<coordinates>(.*?)</coordinates>", pm, re.DOTALL)
        if not coords:
            continue

        partes = coords.group(1).strip().split(",")
        nombre_tag = re.search(r"<name>(.*?)</name>", pm)
        desc_tag   = re.search(r"<description>(.*?)</description>", pm, re.DOTALL)
        desc_raw   = desc_tag.group(1).strip() if desc_tag else ""
        desc_limpia = desc_raw.split("\n")[0].strip()

        if fmt == "mwm":
            # Formato Maps.me: número en <name>, etiqueta en mwm:customName
            custom = re.search(r'<mwm:customName>.*?<mwm:lang[^>]*>(.*?)</mwm:lang>', pm, re.DOTALL)
            nombre = nombre_tag.group(1).strip() if nombre_tag else str(i + 1)
            item   = custom.group(1).strip() if custom else ""
            # Altura en <LookAt><altitude>
            alt    = re.search(r"<altitude>(.*?)</altitude>", pm)
            altura_val = float(alt.group(1)) if alt else 0.0
        else:
            # Formato simple: nombre completo en <name>, altura en <Camera><altitude>
            nombre_completo = nombre_tag.group(1).strip() if nombre_tag else str(i + 1)
            nombre = str(len(puntos) + 1)
            item   = nombre_completo
            # Altura en <Camera><altitude>
            alt    = re.search(r"<Camera>.*?<altitude>(.*?)</altitude>.*?</Camera>", pm, re.DOTALL)
            altura_val = float(alt.group(1)) if alt else 0.0

        puntos.append({
            "N":           len(puntos) + 1,
            "Nombre":      nombre,
            "Item":        item,
            "Descripcion": desc_limpia,
            "Longitud":    round(float(partes[0]), 7),
            "Latitud":     round(float(partes[1]), 7),
            "Altura_m":    round(altura_val, 2) if altura_val > 0 else "",
            "Elevacion_m": "",
        })

    return puntos, fmt


def convertir_utm(puntos: list[dict], epsg: str) -> list[dict]:
    t = Transformer.from_crs("EPSG:4326", epsg, always_xy=True)
    for p in puntos:
        e, n = t.transform(p["Longitud"], p["Latitud"])
        p["Este_UTM"]  = round(e, 2)
        p["Norte_UTM"] = round(n, 2)
    return puntos


def obtener_elevacion(puntos: list[dict], progress_bar, status_text) -> tuple[list[dict], int]:
    BATCH = 100
    BASE  = "https://api.opentopodata.org/v1/srtm90m?locations="
    total = len(puntos)
    ok    = 0

    for i in range(0, total, BATCH):
        lote = puntos[i:i + BATCH]
        locs = "|".join(f"{p['Latitud']},{p['Longitud']}" for p in lote)
        url  = BASE + locs
        try:
            req  = urllib.request.Request(url, headers={"User-Agent": "RedsetelKMZ/1.0"})
            resp = urllib.request.urlopen(req, timeout=20)
            data = json.loads(resp.read())
            for j, result in enumerate(data.get("results", [])):
                elev = result.get("elevation")
                puntos[i + j]["Elevacion_m"] = round(elev, 1) if elev is not None else ""
                ok += 1
        except Exception:
            pass
        progress_bar.progress(min((i + BATCH) / total, 1.0))
        status_text.text(f"Consultando elevación... {min(i + BATCH, total)}/{total} puntos")
        time.sleep(0.5)

    return puntos, ok


def generar_xlsx(puntos: list[dict], zona: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Coordenadas"

    h_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    h_fill  = PatternFill("solid", start_color="0D2650")
    h_align = Alignment(horizontal="center", vertical="center")
    c_align = Alignment(horizontal="center")
    borde   = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )
    fill_par   = PatternFill("solid", start_color="EEF2F8")
    fill_impar = PatternFill("solid", start_color="FFFFFF")

    cols = ["N", "Nombre", "Item", "Descripción",
            "Longitud", "Latitud", "Altura GPS (m)", "Elevación SRTM (m)",
            f"Este UTM ({zona})", f"Norte UTM ({zona})"]

    for col, texto in enumerate(cols, 1):
        c = ws.cell(row=1, column=col, value=texto)
        c.font = h_font; c.fill = h_fill
        c.alignment = h_align; c.border = borde
    ws.row_dimensions[1].height = 22

    for fila, p in enumerate(puntos, 2):
        vals = [p["N"], p["Nombre"], p["Item"], p["Descripcion"],
                p["Longitud"], p["Latitud"], p["Altura_m"], p["Elevacion_m"],
                p["Este_UTM"], p["Norte_UTM"]]
        fill = fill_par if fila % 2 == 0 else fill_impar
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=fila, column=col, value=val)
            c.border = borde; c.alignment = c_align; c.fill = fill
            if col == 5: c.number_format = "0.0000000"
            if col == 6: c.number_format = "0.0000000"
            if col in (7, 8, 9, 10): c.number_format = "#,##0.00"

    anchos = [6, 10, 22, 20, 16, 16, 16, 20, 20, 20]
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Zonas UTM ─────────────────────────────────────────────────────
ZONAS = {
    "17L  — Zona 17 Sur  (EPSG:32717)": ("EPSG:32717", "17L"),
    "18L  — Zona 18 Sur  (EPSG:32718)": ("EPSG:32718", "18L"),
    "19L  — Zona 19 Sur  (EPSG:32719)": ("EPSG:32719", "19L"),
    "17N  — Zona 17 Norte (EPSG:32617)": ("EPSG:32617", "17N"),
    "18N  — Zona 18 Norte (EPSG:32618)": ("EPSG:32618", "18N"),
    "19N  — Zona 19 Norte (EPSG:32619)": ("EPSG:32619", "19N"),
}

FORMATOS = {
    "mwm":    "Maps.me / OruxMaps",
    "simple": "Google Earth / Field Map",
}

# ── HEADER ────────────────────────────────────────────────────────
st.markdown("""
<div class="topbar">
    <div>
        <div class="brand"><span class="red">red</span><span class="white">setel</span></div>
        <div class="subtitle">RED DE SERVICIOS Y TELECOMUNICACIONES PERÚ</div>
    </div>
    <div class="divider"></div>
    <div>
        <div class="tool-title">📡 Extractor KMZ → UTM</div>
        <div class="tool-desc">Coordenadas · Elevación SRTM · Exportación Excel</div>
    </div>
</div>
""", unsafe_allow_html=True)

# ── CONFIGURACIÓN ─────────────────────────────────────────────────
st.markdown('<div class="card card-blue">', unsafe_allow_html=True)
st.markdown('<div class="section-label">⚙ Configuración</div>', unsafe_allow_html=True)
col1, col2 = st.columns([2, 1])
with col1:
    zona_key = st.selectbox("Zona UTM", list(ZONAS.keys()), index=2)
    epsg_sel, nombre_zona = ZONAS[zona_key]
with col2:
    usar_elevacion = st.checkbox("Obtener elevación SRTM", value=True,
                                  help="Consulta elevación real (msnm) desde Open-Topo-Data (NASA SRTM)")
if usar_elevacion:
    st.markdown("""
    <div class="info-box">
        🛰 <strong>Elevación SRTM 90m</strong> — API gratuita Open-Topo-Data · Sin clave requerida ·
        Precisión ~90m horizontal · Fuente: NASA SRTM
    </div>
    """, unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

# ── UPLOAD ────────────────────────────────────────────────────────
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="section-label">📂 Archivo de entrada</div>', unsafe_allow_html=True)
archivo = st.file_uploader("Arrastra tu archivo .kmz aquí", type=["kmz"], label_visibility="collapsed")
st.markdown('</div>', unsafe_allow_html=True)

# ── RESULTADO ─────────────────────────────────────────────────────
if archivo:
    try:
        kml_texto      = leer_kml(archivo.read())
        puntos, fmt    = extraer_puntos(kml_texto)
        fmt_label      = FORMATOS.get(fmt, fmt)

        if not puntos:
            st.warning("No se encontraron puntos en el archivo KMZ.")
        else:
            puntos = convertir_utm(puntos, epsg_sel)

            # Mostrar formato detectado
            st.markdown(f"""
            <div class="info-box">
                ✅ Formato detectado: <strong>{fmt_label}</strong>
                &nbsp;·&nbsp; {len(puntos)} puntos encontrados
            </div>
            """, unsafe_allow_html=True)

            # Elevación SRTM
            elev_ok = 0
            if usar_elevacion:
                st.markdown('<div class="card card-blue">', unsafe_allow_html=True)
                st.markdown('<div class="section-label">🛰 Obteniendo elevación SRTM</div>', unsafe_allow_html=True)
                progress_bar = st.progress(0)
                status_text  = st.empty()
                puntos, elev_ok = obtener_elevacion(puntos, progress_bar, status_text)
                status_text.text(f"✅ Elevación obtenida para {elev_ok}/{len(puntos)} puntos")
                st.markdown('</div>', unsafe_allow_html=True)

            con_item   = sum(1 for p in puntos if p["Item"])
            con_altura = sum(1 for p in puntos if p["Altura_m"] != "")

            st.markdown(f"""
            <div class="metric-row">
                <div class="metric-card">
                    <div class="val">{len(puntos)}</div>
                    <div class="lbl">Puntos totales</div>
                </div>
                <div class="metric-card">
                    <div class="val">{con_item}</div>
                    <div class="lbl">Con item</div>
                </div>
                <div class="metric-card blue">
                    <div class="val">{elev_ok if usar_elevacion else "—"}</div>
                    <div class="lbl">Con elevación</div>
                </div>
                <div class="metric-card blue">
                    <div class="val">{nombre_zona}</div>
                    <div class="lbl">Zona UTM</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown('<div class="section-label">🔍 Vista previa — primeros 10 puntos</div>', unsafe_allow_html=True)
            df = pd.DataFrame(puntos)
            st.dataframe(df.head(10), use_container_width=True, hide_index=True)
            st.markdown('</div>', unsafe_allow_html=True)

            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown('<div class="section-label">⬇ Exportar</div>', unsafe_allow_html=True)
            xlsx_bytes    = generar_xlsx(puntos, nombre_zona)
            nombre_salida = archivo.name.replace(".kmz", f"_UTM_{nombre_zona}.xlsx")
            st.download_button(
                label=f"DESCARGAR EXCEL  ·  {len(puntos)} PUNTOS  ·  ZONA {nombre_zona}",
                data=xlsx_bytes,
                file_name=nombre_salida,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.markdown('</div>', unsafe_allow_html=True)

    except Exception as e:
        st.error(f"Error al procesar el archivo: {e}")
else:
    st.markdown("""
    <div class="card">
        <div class="empty-state">
            <div class="icon">📂</div>
            Sube un archivo <strong>.kmz</strong> para comenzar
        </div>
    </div>
    """, unsafe_allow_html=True)

st.markdown("""
<div class="footer">
    <strong>Redsetel</strong> · Red de Servicios y Telecomunicaciones Perú<br>
    Extracción KMZ · Conversión WGS84 → UTM · Elevación SRTM 90m · Zonas 17–19 N/S
</div>
""", unsafe_allow_html=True)
